"""
Smoothstack Evaluation Week Final Project
"""

import os
from datetime import datetime

from dotenv import load_dotenv
from flask import Flask, flash, redirect, render_template, request, url_for
from openpyxl import load_workbook
from pydantic import ValidationError
from werkzeug.utils import secure_filename

from models import db
from models.summary import Summary
from models.voc import VOC
from utils import helpers
from utils.constants import SUMMARY_SHEET, UPLOADS, VOC_SHEET
from utils.logger import error, info

# load environment variables
load_dotenv()

app = Flask(__name__)

HOST = os.getenv("MYSQL_HOST")
USER = os.getenv("MYSQL_USER")
DATABASE = os.getenv("MYSQL_DATABASE")

app.config['SQLALCHEMY_DATABASE_URI'] = f"mysql+pymysql://{USER}@{HOST}/{DATABASE}"
app.secret_key = "55e36cb88d9251f1bd812ec5242e5ead"


# initialize database connection
db.init_app(app)

# create tables
with app.app_context():
    db.create_all()

# create storage directories if they don't exist
helpers.init_storage()


@app.get('/health')
def check():
    """
    Ping this route to make sure the server is up and running

    Returns:
        str: "OK" if the request successfully hit the server
    """

    return "Ok"


@app.route('/')
def index():
    """
    Home page

    Returns:
        Template: HTML page
    """
    return render_template('home.html')


@app.post("/upload")
def upload():
    """
    POST route to insert Excel data to MySQL

    Returns:
        _type_: _description_
    """

    file = request.files["file"]

    if not file.filename:
        flash("No file uploaded")
        return redirect(url_for('index'))

    # cleanse file path of crazy characters
    filename = secure_filename(file.filename)

    # check if file has already been parsed
    if helpers.has_been_parsed(filename):
        return redirect(url_for('index'))

    file.save(UPLOADS / filename)

    # try to infer month and year from file. If not possible, move to ERROR
    try:
        (month, year) = helpers.get_month_year(filename)
    except (KeyError, ValueError) as err:
        helpers.file_to_errors(filename, err)
        flash("Error: malformed speadsheet file name")
        return redirect(url_for("index"))

    # parse file with openpyxl
    workbook = load_workbook(UPLOADS / filename, data_only=True)

    # check if all three tabs are present (if not, move to ERROR)
    if not helpers.has_required_sheets(workbook.sheetnames):
        flash("Error: malformed speadsheet")

        helpers.file_to_errors(
            filename, f"{filename} doesn't have the expected sheets"
        )

        return redirect(url_for("index"))

    try:
        active_sheet = workbook[SUMMARY_SHEET]
        data_columns = helpers.char_range("A", "F")

        # A well formed summary sheet has relavent data
        # in rows 2 to 14 (for the 12 months)
        for num in range(2, 14):
            values = [active_sheet[f"{c}{num}"].value for c in data_columns]

            row = Summary(
                time_period=values[0],
                calls_offered=values[1],
                abandoned_after_30=values[2],
                fcr=values[3],
                dsat=values[4],
                csat=values[5],
            )

            row.save()

        active_sheet = workbook[VOC_SHEET]

        for col in helpers.char_range("B", "W"):
            cell = active_sheet[f"{col}1"].value

            if not isinstance(cell, datetime):
                flash("Invalid data!")
                helpers.file_to_errors(filename, "Invalid data!")
                return redirect(url_for("index"))

            voc = VOC(
                time_period=cell,
                promoters=active_sheet[f"{col}4"].value,
                passives=active_sheet[f"{col}6"].value,
                detractors=active_sheet[f"{col}8"].value
            )

            voc.save()

        # finished processing, move file to ARCHIVE
        # and return a view with the data
        helpers.file_to_archives(filename)

        return redirect(url_for("results", year=year, month=month))

    except ValidationError as err:
        flash(f"Some of the data in {filename} is invalid!")
        helpers.file_to_errors(filename, err)
        return redirect(url_for("index"))


@app.get("/results/<int:year>/<int:month>")
def results(year, month):
    """
    Display report for a given month and year

    Args:
        year (int): year
        month (int): month
    """
    info(f"Searching summary table for info on {year}-{month:02}")
    summary = Summary.get_entry(year, month)

    if not summary:
        flash(f"No data for {year}-{month:02} found in spreadsheet")
        error(f"No data for {year}-{month:02} found in spreadsheet")
        return redirect(url_for("index"))

    info("Summary info found!")

    info(f"Searching voc table for info on {year}-{month:02}")
    voc = VOC.get_entry(year, month)

    if not voc:
        flash(f"No data for {year}-{month:02} found in spreadsheet")
        error(f"No data for {year}-{month:02} found in spreadsheet")
        return redirect(url_for("index"))

    info("VOC info found!")

    return render_template("results.html", summary=summary.as_dict(), voc=voc.as_dict())


@app.errorhandler(404)
def invalid_route(err):
    """
    Catch-all route for any unrecognized URLs

    Returns:
        Template: Custom 404 page
    """

    error(err)
    return render_template('404.html')
